from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import time
import os

# === CONFIGURATION ===
excel_path = "contacts - Copy.xlsx"
image_path = "C:\\pic.jpeg"
profile_path = os.path.join(os.getcwd(), "whatsapp_profile")  # to persist session
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# === VALIDATE FILES EXIST ===
if not os.path.exists(excel_path):
    print(f"❌ File '{excel_path}' not found.")
    exit()

if not os.path.exists(image_path):
    print(f"❌ Image file '{image_path}' not found.")
    exit()

# === LOAD WORKBOOK ===
wb = load_workbook(excel_path)
ws = wb.active

# === PROMPT: RESET STATUS? ===
reset_choice = input("🔄 Reset previous statuses (y/n)? ").strip().lower()
if reset_choice == "y":
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.cell(row=i, column=2).value = ""
        ws.cell(row=i, column=3).value = ""
    wb.save(excel_path)

# === START SELENIUM DRIVER ===
options = Options()
options.add_argument(f"--user-data-dir={profile_path}")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")
options.add_argument("--remote-debugging-port=9222")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

driver.get("https://web.whatsapp.com")
print("📱 If this is your first time – scan the QR code, otherwise it will continue automatically...")
WebDriverWait(driver, 90).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='3']"))
)

start_time = time.time()  # Timer for total runtime

# === LOAD CONTACTS ===
df = pd.read_excel(excel_path)
names = df["Name"].dropna().tolist()

# === MAIN LOOP ===
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    name = row[0].value
    status_cell = ws.cell(row=i, column=2)
    time_cell = ws.cell(row=i, column=3)

    if status_cell.value:
        print(f"⏭️ Skipping {name} – already processed.")
        continue

    print(f"\n📨 Sending to: {name}")
    try:
        # Exit previous chat
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(0.5)

        # Search contact
        search_box = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='3']"))
        )
        search_box.click()
        search_box.clear()
        search_box.send_keys(Keys.CONTROL, "a")
        search_box.send_keys(Keys.DELETE)
        search_box.send_keys(name)
        time.sleep(1)  # Time for results to load
        search_box.send_keys(Keys.ENTER)  # Select the first result
        time.sleep(1)  # Wait for chat to open

        # Verify correct contact was selected
        try:
            results = WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, '//*[@id="pane-side"]/div[1]/div/div/div')
                )
            )

            contact_found = False
            for result in results:
                try:
                    name_element = result.find_element(By.XPATH, ".//span[@title]")
                    found_name = name_element.get_attribute("title").strip()
                    print(f"👀 Checking: {found_name}")

                    if found_name == name.strip():
                        ActionChains(driver).move_to_element(result).click().perform()
                        contact_found = True
                        print(f"✅ Matching contact found: {found_name}")
                        break

                except Exception as inner_e:
                    print(f"⚠️ Error within result: {inner_e}")
                    continue

            if not contact_found:
                raise Exception("No matching contact found")

        except Exception as e:
            print(f"❌ No contact found named: {name}")
            status_cell.value = "Contact Not Found"
            status_cell.fill = red_fill
            time_cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            time_cell.fill = red_fill
            wb.save(excel_path)
            continue


        # Click attach button (paperclip icon)
        attach_btn = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div/div[1]/button'))
        )
        attach_btn.click()

        # Upload image
        image_input = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']"))
        )
        image_input.send_keys(image_path)

        # Send image
        send_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[3]/div/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div/span'))
        )
        send_button.click()

        # Update status
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_cell.value = "Successful"
        status_cell.fill = green_fill
        time_cell.value = timestamp
        time_cell.fill = green_fill
        print(f"✅ Successfully sent to: {name}")

    except Exception as e:
        msg = str(e)
        print(f"⚠️ Error sending to {name}: {msg}")
        status_cell.value = msg
        status_cell.fill = red_fill
        time_cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_cell.fill = red_fill

    wb.save(excel_path)

# === CLEANUP ===
driver.quit()
wb.save(excel_path)

end_time = time.time()
elapsed = round(end_time - start_time, 2)
print(f"\n🎉 All messages processed! Total time: {elapsed} seconds.")
