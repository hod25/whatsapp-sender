import argparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import time
import os

# === ARGS ===
parser = argparse.ArgumentParser()
parser.add_argument("--profile", required=True, help="תיקיית פרופיל כרום (user-data-dir)")
parser.add_argument("--start", type=int, required=True, help="אינדקס התחלה")
parser.add_argument("--end", type=int, required=True, help="אינדקס סוף")
args = parser.parse_args()

# === CONFIG ===
excel_path = "contacts.xlsx"
image_path = "C:\\pic.jpeg"
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

if not os.path.exists(excel_path) or not os.path.exists(image_path):
    print("❌ קובץ Excel או תמונה חסר.")
    exit()

# === LOAD DATA ===
wb = load_workbook(excel_path)
ws = wb.active
df = pd.read_excel(excel_path)
contacts = df["Name"].dropna().tolist()[args.start:args.end+1]

# === CHROME ===
options = Options()
full_profile_path = os.path.abspath(args.profile)
options.add_argument(f"--user-data-dir={full_profile_path}")
options.add_argument(f"--profile-directory=Profile{args.start}")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--remote-debugging-port=9222")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get("https://web.whatsapp.com")

print("📱 סרוק QR אם נדרש...")
WebDriverWait(driver, 90).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='3']"))
)

start_time = time.time()

# === LOOP ===
for i, name in enumerate(contacts, start=args.start+2):  # +2 עבור שורת כותרת באקסל
    status_cell = ws.cell(row=i, column=2)
    time_cell = ws.cell(row=i, column=3)

    if status_cell.value:
        print(f"⏭️ {name} כבר טופל.")
        continue

    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(0.5)

        search_box = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='3']"))
        )
        search_box.click()
        search_box.clear()
        search_box.send_keys(Keys.CONTROL, "a")
        search_box.send_keys(Keys.DELETE)
        search_box.send_keys(name)
        time.sleep(1)
        search_box.send_keys(Keys.ENTER)
        time.sleep(1)

        results = WebDriverWait(driver, 5).until(
            EC.presence_of_all_elements_located((By.XPATH, '//*[@id="pane-side"]/div[1]/div/div/div'))
        )

        contact_found = False
        for result in results:
            try:
                found_name = result.find_element(By.XPATH, ".//span[@title]").get_attribute("title").strip()
                if found_name == name.strip():
                    ActionChains(driver).move_to_element(result).click().perform()
                    contact_found = True
                    break
            except: continue

        if not contact_found:
            raise Exception("Contact Not Found")

        attach_btn = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div[1]/div/button'))
        )
        attach_btn.click()

        image_input = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']"))
        )
        image_input.send_keys(image_path)

        send_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']"))
        )
        send_button.click()

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_cell.value = "Successful"
        status_cell.fill = green_fill
        time_cell.value = timestamp
        time_cell.fill = green_fill
        print(f"✅ נשלח אל: {name}")

    except Exception as e:
        status_cell.value = str(e)
        status_cell.fill = red_fill
        time_cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_cell.fill = red_fill
        print(f"❌ שגיאה עם {name}: {e}")

    wb.save(excel_path)

# === סיום ===
driver.quit()
wb.save(excel_path)

elapsed = round(time.time() - start_time, 2)
print(f"🎉 סיום {args.profile} | זמן כולל: {elapsed} שניות.")
