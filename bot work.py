import pyautogui
import pandas as pd
import pyperclip
import time
import webbrowser
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# === CONFIGURATION ===
image_path = "C:\\pic.jpeg"
excel_path = "contacts.xlsx"
contacts_image = "contacts_header.png"

# === CHECK FILE EXISTS ===
if not os.path.exists(excel_path):
    print(f"❌ הקובץ '{excel_path}' לא נמצא. ודא שהוא בתיקייה הנכונה.")
    exit()

# === LOAD WORKBOOK ===
wb = load_workbook(excel_path)
ws = wb.active

# === PROMPT: RESET STATUS? ===
reset_choice = input("🔄 האם לאפס סטטוסים קודמים (y/n)? ").strip().lower()

if reset_choice == "y":
    print("🧹 מאפס סטטוסים קיימים...")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.cell(row=i, column=2).value = ""  # סטטוס
        ws.cell(row=i, column=3).value = ""  # זמן
    wb.save(excel_path)
    print("✅ איפוס הסתיים.")
else:
    print("➡️ המשך רגיל – שורות עם סטטוס קודם ידולגו.")

# === OPEN WHATSAPP WEB ===
webbrowser.open("https://web.whatsapp.com")
print("🌐 Opening WhatsApp Web...")
time.sleep(10)

# === LOAD CONTACTS ===
df = pd.read_excel(excel_path)
names = df["Name"].dropna().tolist()

# === STYLE COLORS ===
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# === HELPER FUNCTIONS ===
def wait_and_click(image, timeout=10):
    start = time.time()
    while time.time() - start < timeout:
        try:
            pos = pyautogui.locateCenterOnScreen(image, confidence=0.8)
        except Exception as e:
            print(f"❌ Error locating {image}: {e}")
            return False
        if pos:
            pyautogui.moveTo(pos)
            pyautogui.click()
            return True
        time.sleep(1)
    print(f"❌ Could not find '{image}'")
    return False

def contact_not_found():
    try:
        return pyautogui.locateOnScreen("nofound.png", confidence=0.8) is not None
    except:
        return False

def is_real_contact():
    try:
        return pyautogui.locateOnScreen(contacts_image, confidence=0.8) is not None
    except:
        return False

def clear_search():
    if wait_and_click("searchbar.png"):
        time.sleep(1)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    time.sleep(1.5)

# === MAIN LOOP ===
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    name = row[0].value
    status_cell = ws.cell(row=i, column=2)  # עמודת סטטוס
    time_cell = ws.cell(row=i, column=3)    # עמודת זמן

    if status_cell.value:  # אם כבר טופל – דלג
        print(f"⏭️ Skipping {name} – already has status: {status_cell.value}")
        continue

    print(f"\n🔍 Processing: {name}")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        clear_search()
        pyperclip.copy(name)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(1.5)
        pyautogui.press("enter")
        time.sleep(1.5)

        if contact_not_found():
            raise Exception("Contact not found")

        if not is_real_contact():
            raise Exception("Not an individual contact (probably group)")

        print("📎 Clicking clip icon...")
        if not wait_and_click("clip.png"):
            raise Exception("Clip icon not found")

        time.sleep(2.5)

        print("🖼️ Clicking photo icon...")
        if not wait_and_click("photo.png"):
            raise Exception("Photo icon not found")

        time.sleep(2.5)

        print("📤 Uploading image...")
        pyperclip.copy(image_path)
        pyautogui.hotkey("ctrl", "v")
        pyautogui.press("enter")
        time.sleep(2)

        print("📨 Clicking send...")
        if not wait_and_click("send.png"):
            raise Exception("Send button not found")

        time.sleep(2)
        print(f"✅ Message sent to: {name}")
        status_cell.value = "Successful"
        status_cell.fill = green_fill
        time_cell.value = timestamp
        time_cell.fill = green_fill

    except Exception as e:
        msg = str(e)
        print(f"⚠️ Error sending to {name}: {msg}")
        status_cell.value = msg
        status_cell.fill = red_fill
        time_cell.value = timestamp
        time_cell.fill = red_fill

    wb.save(excel_path)
    time.sleep(2)

# === FINAL SAVE ===
wb.save(excel_path)
print("\n🎉 All messages processed with logs saved!")
